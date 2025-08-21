import io
import base64
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import List, Dict, Any, Optional
import re
import html
import time
from datetime import datetime, timedelta
import socket

# Set default timeout for all socket operations
socket.setdefaulttimeout(30)  # 30 second timeout

class GmailService:
    def __init__(self, gmail_auth):
        self.auth = gmail_auth
        self.creds = self.auth.authenticate()
        self.service = self.auth.get_gmail_service(self.creds)
    
    def get_emails_by_time_range(self, start_datetime: datetime, end_datetime: datetime, label: str = "INBOX") -> List[Dict[str, Any]]:
        """Get emails from a specific time range"""
        try:
            # Build query based on date and time range
            # Gmail API uses 'after' and 'before' for date, and 'rfc822msgid' for specific messages
            # For date range, we can use 'after:YYYY/MM/DD before:YYYY/MM/DD'
            # For time, Gmail API queries are not directly time-based, so we rely on date and then filter locally if needed
            
            query = f"after:{start_datetime.strftime('%Y/%m/%d')} before:{end_datetime.strftime('%Y/%m/%d')}"
            
            if label and label != "ALL":
                query += f" label:{label}"
            
            all_messages = []
            next_page_token = None

            while True:
                try:
                    request_body = {
                        'userId': 'me',
                        'q': query,
                        'maxResults': 500, # Max results per page
                    }
                    if next_page_token:
                        request_body['pageToken'] = next_page_token

                    results = self.service.users().messages().list(**request_body).execute()
                    
                except socket.timeout:
                    print("Request timed out while fetching emails list (time range pagination)")
                    break
                except Exception as e:
                    print(f"API error while fetching emails list (time range pagination): {e}")
                    break
                
                messages = results.get('messages', [])
                for msg in messages:
                    try:
                        full_msg = self.service.users().messages().get(
                            userId='me',
                            id=msg['id'],
                            format='full'
                        ).execute()
                        
                        if self._should_process_email(full_msg, label):
                            all_messages.append(full_msg)
                    except socket.timeout:
                        print(f"Timeout fetching email {msg['id']} (time range pagination)")
                        continue
                    except Exception as e:
                        print(f"Error fetching email {msg['id']} (time range pagination): {e}")
                        continue
                
                next_page_token = results.get('nextPageToken')
                if not next_page_token:
                    break # No more pages
            
            return all_messages
        except Exception as e:
            print(f"Error in get_emails_by_time_range: {e}")
            return []
    
    def get_starred_emails(self, label: str = "STARRED") -> List[Dict[str, Any]]:
        """Get all starred emails"""
        try:
            # Add timeout handling for the API call
            try:
                results = self.service.users().messages().list(
                    userId='me',
                    labelIds=['STARRED', 'UNREAD'],
                    maxResults=20
                ).execute()
            except socket.timeout:
                print("Request timed out while fetching starred emails")
                return []
            except Exception as e:
                print(f"API error while fetching starred emails: {e}")
                return []
            
            messages = results.get('messages', [])
            detailed_messages = []
            
            for msg in messages:
                try:
                    # Add timeout handling for individual email fetch
                    try:
                        full_msg = self.service.users().messages().get(
                            userId='me',
                            id=msg['id'],
                            format='full'
                        ).execute()
                    except socket.timeout:
                        print(f"Timeout fetching starred email {msg['id']}")
                        continue
                    except Exception as e:
                        print(f"Error fetching starred email {msg['id']}: {e}")
                        continue
                    
                    # Check if this email should be processed
                    if self._should_process_email(full_msg, "STARRED"):
                        detailed_messages.append(full_msg)
                except Exception as e:
                    print(f"Error processing starred email {msg['id']}: {e}")
                    continue
            
            return detailed_messages
        except Exception as e:
            print(f"Error in get_starred_emails: {e}")
            return []
    
    
    
    def get_email_summary(self, email_msg: Dict[str, Any]) -> Dict[str, Any]:
        """Extract key information from email for summary"""
        try:
            content = self.extract_email_content(email_msg)
            headers = email_msg.get('payload', {}).get('headers', [])
            
            summary = {
                'id': email_msg['id'],
                'threadId': email_msg.get('threadId'),
                'snippet': email_msg.get('snippet', '')[:100] + '...',
                'content_preview': content[:200] + '...' if len(content) > 200 else content,
                'from': '',
                'to': '',
                'subject': '',
                'date': '',
                'labels': email_msg.get('labelIds', [])
            }
            
            for header in headers:
                name = header.get('name', '').lower()
                value = header.get('value', '')
                
                if name == 'from':
                    match = re.search(r'<(.+)>', value)
                    if match:
                        summary['from'] = match.group(1)
                    else:
                        summary['from'] = value
                elif name == 'to':
                    summary['to'] = value
                elif name == 'subject':
                    summary['subject'] = value
                elif name == 'date':
                    summary['date'] = value
            
            return summary
        except Exception as e:
            print(f"Error in get_email_summary: {e}")
            return {
                'id': email_msg.get('id', 'unknown'),
                'threadId': email_msg.get('threadId'),
                'snippet': 'Error generating summary',
                'content_preview': 'Error generating preview',
                'from': '',
                'to': '',
                'subject': '',
                'date': '',
                'labels': []
            }
    
    def _should_process_email(self, email_msg: Dict[str, Any], current_label: str = "ALL") -> bool:
        """Check if an email should be processed by the agent"""
        try:
            # Skip if already replied by agent
            if self._is_already_replied(email_msg):
                return False
            
            # Skip promotional emails, UNLESS the current filter is specifically "PROMOTIONAL"
            if current_label != "PROMOTIONAL" and self._is_promotional(email_msg):
                return False
            
            # Process all emails that are not replied to by agent and not promotional (or if promotional filter is active)
            return True
        except Exception as e:
            print(f"Error in _should_process_email: {e}")
            return False
    
    def _is_already_replied(self, email_msg: Dict[str, Any]) -> bool:
        """Check if the agent has already replied to this email"""
        thread_id = email_msg.get('threadId')
        if not thread_id:
            return False
        
        try:
            # Add timeout handling for thread fetch
            try:
                thread = self.service.users().threads().get(
                    userId='me', id=thread_id
                ).execute()
            except socket.timeout:
                print(f"Timeout checking thread {thread_id}")
                return False
            except Exception as e:
                print(f"Error checking thread {thread_id}: {e}")
                return False
            
            # Get the user's email address
            try:
                profile = self.service.users().getProfile(userId='me').execute()
                user_email = profile.get('emailAddress', '')
            except Exception as e:
                print(f"Error getting user profile: {e}")
                return False
            
            for msg in thread.get('messages', []):
                payload = msg.get('payload', {})
                headers = payload.get('headers', [])
                
                for header in headers:
                    if header.get('name', '').lower() == 'from':
                        from_email = header.get('value', '')
                        if user_email in from_email:
                            return True
        except Exception as e:
            print(f"Error in _is_already_replied: {e}")
        
        return False
    
    def _is_promotional(self, email_msg: Dict[str, Any]) -> bool:
        """Check if email is promotional"""
        try:
            headers = email_msg.get('payload', {}).get('headers', [])
            
            for header in headers:
                name = header.get('name', '').lower()
                value = header.get('value', '').lower()
                
                if name == 'list-unsubscribe':
                    return True
                if name == 'precedence' and 'bulk' in value:
                    return True
                if name == 'x-category' and ('promotion' in value or 'newsletter' in value):
                    return True
            
            # Check subject for promotional keywords
            subject = ''
            for header in headers:
                if header.get('name', '').lower() == 'subject':
                    subject = header.get('value', '').lower()
                    break
            
            promotional_keywords = ['sale', 'discount', 'offer', 'promotion', 'newsletter', 'marketing', 'deal', 'coupon', 'save', 'limited time']
            if any(keyword in subject for keyword in promotional_keywords):
                return True
            
            return False
        except Exception as e:
            print(f"Error in _is_promotional: {e}")
            return False
    
    def extract_email_content(self, email_msg: Dict[str, Any]) -> str:
        """Extract and clean email content"""
        try:
            payload = email_msg.get('payload', {})
            
            # Get subject
            subject = ''
            headers = payload.get('headers', [])
            for header in headers:
                if header.get('name', '').lower() == 'subject':
                    subject = header.get('value', '')
                    break
            
            # Get body content
            body = self._extract_body(payload)
            
            # Clean and format content
            if body:
                # Remove HTML tags
                body = re.sub(r'<[^>]+>', ' ', body)
                # Decode HTML entities
                body = html.unescape(body)
                # Normalize whitespace
                body = re.sub(r'\s+', ' ', body).strip()
            
            return f"Subject: {subject}\n\n{body}"
        except Exception as e:
            print(f"Error in extract_email_content: {e}")
            return "Error extracting email content"
    
    def _extract_body(self, payload: Dict[str, Any]) -> str:
        """Extract email body from payload"""
        try:
            body = ''
            
            # Check if body data is available
            if 'body' in payload and 'data' in payload['body']:
                try:
                    body_data = payload['body']['data']
                    body = base64.urlsafe_b64decode(body_data).decode('utf-8')
                    return body
                except Exception:
                    pass
            
            # Check parts for body content
            if 'parts' in payload:
                for part in payload['parts']:
                    if part.get('mimeType') == 'text/plain':
                        if 'data' in part['body']:
                            try:
                                body_data = part['body']['data']
                                body = base64.urlsafe_b64decode(body_data).decode('utf-8')
                                break
                            except Exception:
                                continue
                    # Recursively check nested parts
                    elif 'parts' in part:
                        nested_body = self._extract_body(part)
                        if nested_body:
                            body = nested_body
                            break
            
            return body
        except Exception as e:
            print(f"Error in _extract_body: {e}")
            return ""
    
    def extract_attachments(self, email_msg: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract attachments from an email message."""
        attachments = []
        payload = email_msg.get('payload', {})
        
        if 'parts' in payload:
            for part in payload['parts']:
                filename = part.get('filename')
                mime_type = part.get('mimeType')
                body = part.get('body')

                if filename and body: # It's an attachment if it has a filename and a body
                    attachment_data = {
                        'filename': filename,
                        'mimeType': mime_type,
                        'data': None,
                        'attachmentId': None
                    }
                    
                    if 'data' in body:
                        attachment_data['data'] = body['data']
                    elif 'attachmentId' in body: # For large attachments
                        attachment_data['attachmentId'] = body['attachmentId']
                        # You would typically fetch the actual attachment content here using
                        # self.service.users().messages().attachments().get(userId='me', messageId=email_msg['id'], id=body['attachmentId']).execute()
                        # For now, we'll just store the ID.
                        print(f"DEBUG: Large attachment found: {filename}. Attachment ID: {body['attachmentId']}")

                    attachments.append(attachment_data)
                
                # Recursively check nested parts for attachments
                if 'parts' in part:
                    attachments.extend(self.extract_attachments({'payload': part})) # Pass nested payload

        return attachments
    
    def read_attachment_content(self, attachment: Dict[str, Any]) -> str:
        """Reads the content of an attachment based on its MIME type."""
        content = ""
        mime_type = attachment.get('mimeType')
        data = attachment.get('data')

        if not data:
            # For large attachments, we would need to fetch content using attachmentId
            # This is a placeholder for now.
            print(f"DEBUG: No direct data for attachment {attachment.get('filename')}. Skipping content reading.")
            return ""

        try:
            decoded_data = base64.urlsafe_b64decode(data)

            if mime_type == 'application/pdf':
                try:
                    from PyPDF2 import PdfReader
                    pdf_reader = PdfReader(io.BytesIO(decoded_data))
                    for page in pdf_reader.pages:
                        content += page.extract_text() + "\n"
                except Exception as e:
                    print(f"Error reading PDF attachment: {e}")
                    content = f"Could not read PDF: {e}"
            elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document': # .docx
                try:
                    from docx import Document
                    document = Document(io.BytesIO(decoded_data))
                    for para in document.paragraphs:
                        content += para.text + "\n"
                except Exception as e:
                    print(f"Error reading DOCX attachment: {e}")
                    content = f"Could not read DOCX: {e}"
            elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': # .xlsx
                try:
                    from openpyxl import load_workbook
                    workbook = load_workbook(io.BytesIO(decoded_data))
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value is not None:
                                    content += str(cell.value) + "\t"
                            content += "\n"
                except Exception as e:
                    print(f"Error reading XLSX attachment: {e}")
                    content = f"Could not read XLSX: {e}"
            elif mime_type.startswith('text/'):
                content = decoded_data.decode('utf-8', errors='ignore')
            else:
                content = f"Unsupported attachment type: {mime_type}"

        except Exception as e:
            print(f"Error decoding or reading attachment content: {e}")
            content = f"Error processing attachment: {e}"

        return content
    
    def create_draft_reply(self, original_msg: Dict[str, Any], reply_content: str) -> bool:
        """Create a draft reply to an email"""
        try:
            # Get original message details
            thread_id = original_msg.get('threadId')
            headers = original_msg.get('payload', {}).get('headers', [])
            
            # Extract recipient and subject
            to_email = ''
            subject = ''
            message_id = ''
            
            for header in headers:
                name = header.get('name', '').lower()
                value = header.get('value', '')
                
                if name == 'from':
                    match = re.search(r'<(.+)>', value)
                    if match:
                        to_email = match.group(1)
                    else:
                        to_email = value
                elif name == 'subject':
                    subject = value
                    if not subject.lower().startswith('re:'):
                        subject = f"Re: {subject}"
                elif name == 'message-id':
                    message_id = value
            
            # Create MIME message
            message = MIMEMultipart()
            message['to'] = to_email
            message['subject'] = subject
            if message_id:
                message['In-Reply-To'] = message_id
                message['References'] = message_id
            
            # Add reply content
            message.attach(MIMEText(reply_content, 'plain'))
            
            # Encode message
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
            
            # Create draft with timeout handling
            try:
                draft_body = {
                    'message': {
                        'raw': raw_message,
                        'threadId': thread_id
                    }
                }
                
                self.service.users().drafts().create(
                    userId='me',
                    body=draft_body
                ).execute()
            except socket.timeout:
                print("Timeout creating draft")
                return False
            except Exception as e:
                print(f"Error creating draft: {e}")
                return False
            
            # Remove star to prevent reprocessing (if starred)
            if 'STARRED' in original_msg.get('labelIds', []):
                self._remove_star(original_msg['id'])
            
            return True
            
        except Exception as e:
            print(f"Error in create_draft_reply: {e}")
            return False
    
    def _remove_star(self, message_id: str):
        """Remove star from processed email"""
        try:
            # Add timeout handling for star removal
            try:
                self.service.users().messages().modify(
                    userId='me',
                    id=message_id,
                    body={'removeLabelIds': ['STARRED']}
                ).execute()
            except socket.timeout:
                print(f"Timeout removing star from message {message_id}")
            except Exception as e:
                print(f"Error removing star from message {message_id}: {e}")
        except Exception as e:
            print(f"Error in _remove_star: {e}")